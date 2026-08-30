"""The compliance matrix, and its round trip through a spreadsheet.

Requirement extraction is the highest-leverage step in a proposal and the one a
model gets imperfectly right. A requirement missed here is a requirement nobody
writes to, and no amount of good drafting downstream recovers it. So the matrix
is the one artifact built to be corrected by a human and read back.

The round trip is deliberately a spreadsheet. Not because XLSX is elegant, but
because the proposal manager who has to check three hundred extracted
requirements already lives in Excel, and a tool that makes that person learn a
new interface to do the most important review in the process will not be used.

`import_workbook` reports what changed rather than silently absorbing it: edits,
added rows, and rows a human flagged as mis-extracted all come back named, so
the correction itself is auditable.
"""

from __future__ import annotations

import json
import re
from dataclasses import asdict, dataclass, field, fields
from pathlib import Path

# Column order is the schema in reference/compliance-schema.md, plus the three
# columns that exist only for the human round trip.
COLUMNS = [
    ("id", "ID"),
    ("source", "Source"),
    ("requirement", "Requirement"),
    ("type", "Type"),
    ("volume", "Volume"),
    ("section", "Section"),
    ("eval_criterion", "Eval criterion"),
    ("page_limit", "Page limit"),
    ("owner", "Owner"),
    ("status", "Status"),
    ("reviewer_note", "Reviewer note"),
    ("flag", "Flag"),
]

FLAGS = {"", "ok", "mis-extracted", "duplicate", "added", "not-a-requirement"}

STATUSES = {"open", "drafted", "reviewed", "complete"}


@dataclass
class Requirement:
    id: str
    requirement: str
    source: str = ""
    type: str = "submission"
    volume: str = ""
    section: str = ""
    eval_criterion: str = ""
    page_limit: str = ""
    owner: str = ""
    status: str = "open"
    reviewer_note: str = ""
    flag: str = ""

    def as_dict(self) -> dict:
        return asdict(self)

    @property
    def live(self) -> bool:
        """Rows a human marked as junk do not count against coverage."""
        return self.flag not in {"mis-extracted", "duplicate", "not-a-requirement"}


def _clean(value) -> str:
    if value is None:
        return ""
    return " ".join(str(value).split())


def from_markdown(text: str) -> list[Requirement]:
    """Parse the markdown table an extraction agent produces.

    Tolerant by design: agents vary the column order and sometimes add columns,
    so rows are mapped by header name rather than position, and unknown columns
    are ignored instead of shifting everything.
    """
    known = {label.lower(): key for key, label in COLUMNS}
    valid = {f.name for f in fields(Requirement)}
    rows: list[Requirement] = []
    header: list[str] | None = None

    for line in text.splitlines():
        line = line.strip()
        if not line.startswith("|"):
            header = None
            continue
        cells = [c.strip() for c in line.strip("|").split("|")]
        if all(set(c) <= {"-", ":", " "} for c in cells if c):
            continue  # the ---|--- separator row
        if header is None:
            header = [known.get(c.lower(), "") for c in cells]
            continue

        data = {}
        for key, cell in zip(header, cells):
            if key and key in valid:
                data[key] = _clean(cell)
        if not data.get("id") or not data.get("requirement"):
            continue
        rows.append(Requirement(**data))

    return rows


def to_markdown(rows: list[Requirement]) -> str:
    head = "| " + " | ".join(label for _, label in COLUMNS) + " |"
    rule = "| " + " | ".join("---" for _ in COLUMNS) + " |"
    body = [
        "| " + " | ".join(_clean(getattr(r, key)) for key, _ in COLUMNS) + " |"
        for r in rows
    ]
    return "\n".join([head, rule, *body])


def save_json(rows: list[Requirement], path: Path | str) -> Path:
    path = Path(path)
    path.write_text(
        json.dumps([r.as_dict() for r in rows], indent=2), encoding="utf-8"
    )
    return path


def load_json(path: Path | str) -> list[Requirement]:
    data = json.loads(Path(path).read_text(encoding="utf-8"))
    valid = {f.name for f in fields(Requirement)}
    return [Requirement(**{k: v for k, v in row.items() if k in valid}) for row in data]


def load(path: Path | str) -> list[Requirement]:
    """Read a matrix from .json, .xlsx, or a markdown table."""
    path = Path(path)
    if path.suffix.lower() == ".json":
        return load_json(path)
    if path.suffix.lower() in {".xlsx", ".xlsm"}:
        return import_workbook(path)[0]
    return from_markdown(path.read_text(encoding="utf-8"))


# --- the spreadsheet round trip ------------------------------------------

def export_workbook(rows: list[Requirement], path: Path | str) -> Path:
    """Write the matrix as an editable workbook with guidance and validation."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.datavalidation import DataValidation
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError(
            "Exporting a workbook needs openpyxl. Install it with: pip install openpyxl"
        ) from exc

    path = Path(path)
    book = Workbook()
    sheet = book.active
    sheet.title = "Compliance matrix"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1D6B4C")

    sheet.append([label for _, label in COLUMNS])
    for cell in sheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center")
    sheet.freeze_panes = "A2"

    for row in rows:
        sheet.append([_clean(getattr(row, key)) for key, _ in COLUMNS])

    widths = {"ID": 12, "Source": 18, "Requirement": 70, "Type": 14, "Volume": 10,
              "Section": 12, "Eval criterion": 16, "Page limit": 11, "Owner": 16,
              "Status": 12, "Reviewer note": 40, "Flag": 18}
    for index, (_, label) in enumerate(COLUMNS, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = widths.get(label, 16)
    for row_cells in sheet.iter_rows(min_row=2):
        row_cells[2].alignment = Alignment(wrap_text=True, vertical="top")
        row_cells[10].alignment = Alignment(wrap_text=True, vertical="top")

    last = max(sheet.max_row, 2)
    flag_column = get_column_letter(len(COLUMNS))
    status_column = get_column_letter(len(COLUMNS) - 2)

    flag_rule = DataValidation(
        type="list",
        formula1='"ok,mis-extracted,duplicate,added,not-a-requirement"',
        allow_blank=True,
    )
    sheet.add_data_validation(flag_rule)
    flag_rule.add(f"{flag_column}2:{flag_column}{last + 200}")

    status_rule = DataValidation(
        type="list", formula1='"open,drafted,reviewed,complete"', allow_blank=True
    )
    sheet.add_data_validation(status_rule)
    status_rule.add(f"{status_column}2:{status_column}{last + 200}")

    notes = book.create_sheet("How to use this")
    for line in [
        ["Correcting the compliance matrix"],
        [""],
        ["This workbook is the one place a human is expected to correct the"],
        ["machine. Extraction is imperfect, and a requirement missed here is a"],
        ["requirement nobody writes to."],
        [""],
        ["Edit any cell. Add rows at the bottom for requirements the extraction"],
        ["missed, and set their Flag to 'added'."],
        [""],
        ["Flag values:"],
        ["  ok                 correct as extracted"],
        ["  mis-extracted      not actually a requirement, or wrongly captured"],
        ["  duplicate          the same obligation as another row"],
        ["  added              a requirement you added by hand"],
        ["  not-a-requirement  narrative text the extractor mistook for an obligation"],
        [""],
        ["Rows flagged mis-extracted, duplicate, or not-a-requirement are"],
        ["excluded from coverage checks. Nothing is deleted — the record of what"],
        ["was extracted and what a human rejected stays intact."],
        [""],
        ["Do not renumber the ID column. IDs are referenced from the outline,"],
        ["the draft, and every review comment."],
        [""],
        ["Read it back with:  colorteam matrix import <this file> -o matrix.json"],
    ]:
        notes.append(line)
    notes.column_dimensions["A"].width = 76
    notes["A1"].font = Font(bold=True, size=14)

    book.save(path)
    return path


def import_workbook(path: Path | str) -> tuple[list[Requirement], dict]:
    """Read a corrected workbook back, returning the rows and what changed."""
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError(
            "Reading a workbook needs openpyxl. Install it with: pip install openpyxl"
        ) from exc

    book = load_workbook(Path(path), data_only=True)
    sheet = book["Compliance matrix"] if "Compliance matrix" in book.sheetnames else book.active

    header_cells = [_clean(c.value) for c in sheet[1]]
    label_to_key = {label.lower(): key for key, label in COLUMNS}
    header = [label_to_key.get(h.lower(), "") for h in header_cells]
    valid = {f.name for f in fields(Requirement)}

    rows: list[Requirement] = []
    for excel_row in sheet.iter_rows(min_row=2, values_only=True):
        data = {}
        for key, value in zip(header, excel_row):
            if key and key in valid:
                data[key] = _clean(value)
        if not data.get("id") or not data.get("requirement"):
            continue
        rows.append(Requirement(**data))

    report = {
        "total": len(rows),
        "live": sum(1 for r in rows if r.live),
        "added_by_hand": [r.id for r in rows if r.flag == "added"],
        "rejected": [
            {"id": r.id, "flag": r.flag, "note": r.reviewer_note}
            for r in rows
            if not r.live
        ],
        "with_notes": [r.id for r in rows if r.reviewer_note],
    }
    return rows, report


def diff(before: list[Requirement], after: list[Requirement]) -> dict:
    """What a human changed between export and import."""
    old = {r.id: r for r in before}
    new = {r.id: r for r in after}

    edited = []
    for ident, row in new.items():
        previous = old.get(ident)
        if previous is None:
            continue
        changed = [
            key for key, _ in COLUMNS
            if key not in {"reviewer_note", "flag"}
            and _clean(getattr(previous, key)) != _clean(getattr(row, key))
        ]
        if changed:
            edited.append({"id": ident, "fields": changed})

    return {
        "added": sorted(set(new) - set(old)),
        "removed": sorted(set(old) - set(new)),
        "edited": edited,
        "rejected": [r.id for r in after if not r.live],
    }
